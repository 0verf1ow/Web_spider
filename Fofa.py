# /bin/bash/python3
# _*_coding:utf-8_*_
"""
更新时间：2021.4.3
"""
import threading
from datetime import datetime as datetime_datetime
from requests import get as r_get
from re import search as re_search, findall as re_findall
from base64 import b64encode as base64_b64encode
from time import sleep as time_sleep
from openpyxl import Workbook
from openpyxl.styles import Alignment
from urllib.parse import quote as urllib_parse_quote
import math
import json


class FoFaSpider():
    """初始化配置"""

    def __init__(self, search_keyword, Auth, start_page, end_page, thread, proxies):
        self.proxies = proxies
        self.mHeaders = {}
        self.mHeaders[
            "User-Agent"] = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/84 Safari/537.36"
        self.mHeaders["Authorization"] = Auth
        self.mHeaders["Cookie"] = "fofa_token={}".format(Auth)
        self.url = 'https://api.fofa.so/v1/search?&full=false&pn={}&ps=20&qbase64=' + urllib_parse_quote(
            base64_b64encode(search_keyword.encode()).decode())
        self.filename = datetime_datetime.now().strftime('%Y-%m-%d_%H_%M_%S') + '.xlsx'  # 保存的文件名
        self.start_page = start_page  # 开始爬取的页码，默认是1
        self.end_page = end_page  # 结束爬取的页码
        self.thread = thread  # 多线程数量
        self.error = 0  # 累计错误次数，达 5 次则停止运行
        self.data = []  # 用于存储爬取的数据,一条结果是一个小元组

    """爬虫主线程"""

    def run(self):
        self.max_page = self.get_pages()  # 爬取的最大页码
        self.end_page = self.max_page if end_page == 0 else end_page  # 获取要爬取的页数
        print("[+] 查询结果一共是 {} 页，本次爬取 {} 页保存".format(self.max_page, self.end_page))

        thread_list = []  # 任务列表
        url_list = []  # 密码列表
        for i in range(self.start_page, self.end_page + 1):
            url = self.url.format(i)
            url_list.append((url, i))

        self.threadmax = threading.BoundedSemaphore(self.thread)  # 限制线程的最大数量

        # 创建队列
        for u in url_list:
            t = threading.Thread(target=self.spider, args=(u[0], u[1]))
            thread_list.append(t)

        for thread in thread_list:
            self.threadmax.acquire()
            thread.start()

        for thread in thread_list:
            thread.join()

        self.data = set(self.data)  # 去重
        self.xls()

    """HTTP网络抓取"""

    def spider(self, url, n):
        try:
            r = r_get(url=url, headers=self.mHeaders, proxies=self.proxies,timeout=5)
            result = json.loads(r.text)
            if result['code'] == 0:
                if len(result['data']['assets']) == 0:
                    self.xls()
                else:
                    print("[+] 正在爬取第 {} 页,正常运行".format(n, ))
            elif result['code'] == 820004:
                print("[!] 正在爬取第 {} 页,资源访问权限不足，当前账户不允许爬到此页".format(n))
                self.xls(True)
                exit()
            data = result['data']['assets']

            for d in data:
                link = d['link']
                ip = d['ip']
                title = d['title']
                header = d['header']
                code = re_search(r'HTTP/1.* (\d+)', d['header'])
                code = code.group(1) if code else ''
                port = d['port']
                protocol = d['protocol']
                server = d['server']
                self.data.append((link, ip, port, protocol, code, title, server, header))
            self.threadmax.release()  # 释放信号量，可用信号量加一
            return 0

        except Exception as e:
            print("[!]错误信息：" + str(e))
            for s in range(1, 31).__reversed__():
                print("\r[!] 状态码异常, 正在休眠 {} 秒后重试".format(s), end='', flush=True)
                time_sleep(1)
            print("\n[*] 休眠结束，继续开始爬取")
            self.error += 1
            if self.error == 3:
                print('[!]失败次数太多，已停止运行，请检查后运行')
                self.xls(True)
            self.spider(url, n)
        self.threadmax.release()  # 释放信号量，可用信号量加一
        return 0

    """爬取结果写入excel表格"""

    def xls(self, error=False):
        wb = Workbook()
        ws = wb.active

        # 表格宽度
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 30
        ws.column_dimensions['G'].width = 30
        ws.column_dimensions['H'].width = 50

        # 表头
        ws['A1'] = '链接'
        ws['B1'] = 'IP'
        ws['C1'] = '端口'
        ws['D1'] = '协议'
        ws['E1'] = '状态码'
        ws['F1'] = '标题'
        ws['G1'] = '中间件'
        ws['H1'] = '返回头'

        for d in self.data:
            ws.append([d[0], d[1], d[2], d[3], d[4], d[5], d[6], d[7].strip()])

        nrows = ws.max_row  # 获得行数
        ncols = ws.max_column
        for i in range(nrows):
            for j in range(1, ncols - 1):
                ws.cell(row=i + 1, column=j + 1).alignment = Alignment(horizontal='center', vertical='center')
        wb.save(self.filename)

        if error == True:
            self.xls()
            print("[!] 爬取异常结束, 已经爬取的结果保存至运行目录下的 {} ".format(self.filename))
            exit()
        else:
            print("[*] 爬取任务结束, 爬取结果已经保存至运行目录下的 {} ".format(self.filename))

    """获取总页数"""

    def get_pages(self):
        url = 'https://fofa.so/result?page=1&qbase64=' + urllib_parse_quote(
            base64_b64encode(search_keyword.encode()).decode())
        r = r_get(url, headers=self.mHeaders, proxies=self.proxies)
        pages = re_findall('class="number">(\d*)</li></ul><button type="button" class="btn-next">', r.text)
        end_page = pages[0] if pages else 1
        return math.ceil(int(end_page) / 2)


if __name__ == '__main__':
    ######################################## 配置区域 ##########################################
    # 搜索的语法
    search_keyword = 'fid="giaSqkgiXSKzo5S7C4a0og=="'
    ##
    # HTTP包头部中的Authorization，或者Cookie中的fofa_token,用于验证身份
    Auth = "eyJhbGciOiJIUzUxMiIsImtpZCI6Ik5XWTVZakF4TVRkalltSTJNRFZsWXpRM05EWXdaakF3TURVMlkyWTNZemd3TUdRd1pUTmpZUT09IiwidHlwIjoiSldUIn0.eyJpZCI6NTQ1MjQsIm1pZCI6MTAwMDM2MDY1LCJ1c2VybmFtZSI6IuS7iuaZmuaJk-iAgeiZjiIsImV4cCI6MTYzNjY1NzkwMH0.4H9KxN5j5pei4uRKisQGCyzqcdx_ziyqtrkzj0LUeoDF09dO6n_yukauSPK25mjLOqVcFteDyBbSJiRAnvgFYQ"
    ##
    # 代理
    open_proxy = 0  # 代理开关
    proxies = {"https": "http://127.0.0.1:8080"}  # 代理地址，支持 socks、http(s)代理
    ##
    # 开始爬的页码,默认是1
    start_page = 1
    ##
    # 爬到多少页,不改则爬出所有页
    end_page = 0
    ##
    # 多线程数量
    thread = 10
    ##################################### 配置区域结束 ##########################################

    # 创建爬虫
    f = FoFaSpider(search_keyword, Auth, start_page, end_page, thread, proxies if open_proxy else None)
    f.run()  # 启动爬虫
